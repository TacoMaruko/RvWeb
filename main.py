from flask import Flask, render_template, request

app = Flask(__name__)


@app.route("/", methods=['get', 'post'])
def index():
    print(f"request 的 attribute method:{request.method}")
    print(f"request 的 attribute mimetype:{request.mimetype_params}")
    return render_template('index.html')


@app.route('/name/<username>')
def show_user(username):
    return f"<h2>你的名稱是：<strong>{username}</strong></h2>"

from openpyxl import load_workbook
@app.route("/login", methods=['get', 'post'])
def login():
    if request.method == "POST":
        print("由表單傳送來的是單行")
        print(request.form)
        print("=== 底下分成二行，較清楚 ===")
        print(f"email:{request.form['email']}")
        print(f"password:{request.form['passwd']}")
        wb = load_workbook('static/others/ePW-1.xlsx')
        login_sheet = wb['工作表1']
        for row in login_sheet.iter_rows():
            print()
        # print(wb.工作表1)
    return render_template('login.html')